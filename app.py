from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import io
import openpyxl 
from fpdf import FPDF

app = Flask(__name__)
app.config['SECRET_KEY'] = 'church-teens-secret-key-v2'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///church.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- Models ---
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)

class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    date = db.Column(db.String(50), nullable=False)
    description = db.Column(db.Text, nullable=True)
    attendees = db.relationship('Attendee', backref='event', lazy=True, cascade="all, delete")

class Attendee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    phone = db.Column(db.String(50), nullable=True)
    parent_phone = db.Column(db.String(50), nullable=True)
    email = db.Column(db.String(120), nullable=True)
    address = db.Column(db.String(200), nullable=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Dashboard & Auth ---
@app.route('/')
@login_required
def dashboard():
    events = Event.query.order_by(Event.date).all()
    return render_template('dashboard.html', events=events)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- Event Management ---
@app.route('/create_event', methods=['POST'])
@login_required
def create_event():
    title = request.form.get('title')
    date = request.form.get('date')
    description = request.form.get('description')
    db.session.add(Event(title=title, date=date, description=description))
    db.session.commit()
    flash('Event created successfully!')
    return redirect(url_for('dashboard'))

@app.route('/delete_event/<int:event_id>')
@login_required
def delete_event(event_id):
    event = Event.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    flash('Event deleted.')
    return redirect(url_for('dashboard'))

@app.route('/edit_event/<int:event_id>', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    event = Event.query.get_or_404(event_id)
    if request.method == 'POST':
        event.title = request.form.get('title')
        event.date = request.form.get('date')
        event.description = request.form.get('description')
        db.session.commit()
        flash('Event updated.')
        return redirect(url_for('dashboard'))
    return render_template('edit_event.html', event=event)

@app.route('/event/<int:event_id>', methods=['GET', 'POST'])
@login_required
def event_detail(event_id):
    event = Event.query.get_or_404(event_id)
    if request.method == 'POST':
        # Add Attendee
        db.session.add(Attendee(
            name=request.form.get('name'),
            phone=request.form.get('phone'),
            parent_phone=request.form.get('parent_phone'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            event_id=event.id
        ))
        db.session.commit()
        return redirect(url_for('event_detail', event_id=event.id))
    return render_template('event_detail.html', event=event)

# --- Attendee Management ---
@app.route('/delete_attendee/<int:id>')
@login_required
def delete_attendee(id):
    attendee = Attendee.query.get_or_404(id)
    event_id = attendee.event_id
    db.session.delete(attendee)
    db.session.commit()
    flash('Attendee removed.')
    return redirect(url_for('event_detail', event_id=event_id))

@app.route('/edit_attendee/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_attendee(id):
    attendee = Attendee.query.get_or_404(id)
    if request.method == 'POST':
        attendee.name = request.form.get('name')
        attendee.phone = request.form.get('phone')
        attendee.parent_phone = request.form.get('parent_phone')
        attendee.email = request.form.get('email')
        attendee.address = request.form.get('address')
        db.session.commit()
        flash('Attendee updated.')
        return redirect(url_for('event_detail', event_id=attendee.event_id))
    return render_template('edit_attendee.html', attendee=attendee)

# --- Admin Management ---
@app.route('/register_admin', methods=['GET', 'POST'])
@login_required
def register_admin():
    users = User.query.all()
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if User.query.filter_by(username=username).first():
            flash('Username already exists.')
        else:
            hashed_pw = generate_password_hash(password, method='scrypt')
            db.session.add(User(username=username, password=hashed_pw))
            db.session.commit()
            flash('New Admin created successfully.')
            return redirect(url_for('register_admin'))
    return render_template('register_admin.html', users=users)

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_user.password = generate_password_hash(request.form.get('password'), method='scrypt')
        db.session.commit()
        flash('Password updated.')
        return redirect(url_for('dashboard'))
    return render_template('change_password.html')

# --- Exports ---
@app.route('/export/excel/<int:event_id>')
@login_required
def export_excel(event_id):
    event = Event.query.get_or_404(event_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendees"
    
    # Header
    headers = ['Name', 'Phone', 'Parent Phone', 'Email', 'Address']
    ws.append(headers)
    
    for a in event.attendees:
        ws.append([a.name, a.phone, a.parent_phone, a.email, a.address])
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, download_name=f"{event.title}_attendees.xlsx", as_attachment=True)

@app.route('/export/pdf/<int:event_id>')
@login_required
def export_pdf(event_id):
    event = Event.query.get_or_404(event_id)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    pdf.cell(200, 10, txt=f"Event: {event.title}", ln=True, align='C')
    pdf.cell(200, 10, txt=f"Date: {event.date}", ln=True, align='C')
    pdf.ln(10)
    
    # Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 10, 'Name', 1)
    pdf.cell(35, 10, 'Phone', 1)
    pdf.cell(35, 10, 'Parent Ph', 1)
    pdf.cell(45, 10, 'Email', 1)
    pdf.ln()
    
    # Rows
    pdf.set_font("Arial", size=10)
    for a in event.attendees:
        # Simple cell clipping
        name = (a.name[:18] + '..') if len(a.name) > 20 else a.name
        email = (a.email[:23] + '..') if a.email and len(a.email) > 25 else (a.email or "")
        
        pdf.cell(40, 10, name, 1)
        pdf.cell(35, 10, a.phone or "", 1)
        pdf.cell(35, 10, a.parent_phone or "", 1)
        pdf.cell(45, 10, email, 1)
        pdf.ln()
        
    out = io.BytesIO(pdf.output(dest='S').encode('latin1')) # FPDF output to bytes
    return send_file(out, download_name=f"{event.title}_attendees.pdf", as_attachment=True)

def create_default_admin():
    if not User.query.filter_by(username='admin').first():
        hashed_pw = generate_password_hash('admin', method='scrypt')
        db.session.add(User(username='admin', password=hashed_pw))
        db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        create_default_admin()

    app.run(debug=False, host="0.0.0.0")
